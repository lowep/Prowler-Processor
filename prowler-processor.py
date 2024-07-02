import json
import os
import argparse
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# Configuration options
RESOURCE_TYPES = []  # Leave empty to include all, or specify types to include
EXCLUDE_CHECK_TYPES = []  # Check types to exclude
MAX_FINDINGS_PER_CHECK = 10  # Maximum number of findings to include per check type

def load_json_file(file_path):
    with open(file_path, 'r') as file:
        return json.load(file)

def process_prowler_scans(file_paths, severity_filter):
    all_findings = defaultdict(list)
    
    for file_path in file_paths:
        print(f"Processing file: {file_path}")  # Added print statement
        data = load_json_file(file_path)
        for finding in data:
            # Apply filters
            severity = finding.get('Severity', '')
            if isinstance(severity, dict):
                severity = severity.get('Label') or next(iter(severity.values()), '')
            severity = str(severity).upper()
            
            if severity[0] not in severity_filter:  # Check only the first letter
                continue
            
            if RESOURCE_TYPES and not any(r.get('Type') in RESOURCE_TYPES for r in finding.get('Resources', [])):
                continue
            
            if any(check_type in finding.get('Types', []) for check_type in EXCLUDE_CHECK_TYPES):
                continue
            
            key = (
                tuple(finding.get('Types', [])),
                finding.get('Title', ''),
                severity
            )
            all_findings[key].append(finding)
    
    # Group similar findings and apply MAX_FINDINGS_PER_CHECK
    grouped_findings = []
    for key, findings in all_findings.items():
        grouped_findings.extend(findings[:MAX_FINDINGS_PER_CHECK])
    
    return grouped_findings

def save_excel_file(data, output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Findings"

    fields = ['Severity', 'Title', 'Description', 'Types', 'Resources', 'SchemaVersion', 'ProductFields']
    
    # Header styling
    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    header_font = Font(bold=True)
    header_border = Border(bottom=Side(style='medium'))
    
    for col, header in enumerate(fields, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = header_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    severity_order = {'CRITICAL': 0, 'HIGH': 1, 'MEDIUM': 2, 'LOW': 3}
    
    def get_severity(x):
        severity = x.get('Severity', '')
        if isinstance(severity, dict):
            severity = severity.get('Label') or next(iter(severity.values()), '')
        return severity_order.get(str(severity).upper(), 4)
    
    sorted_data = sorted(data, key=get_severity)

    # Alternating row colors
    alt_row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    for row, finding in enumerate(sorted_data, start=2):
        severity = finding.get('Severity', '')
        if isinstance(severity, dict):
            severity = severity.get('Label') or next(iter(severity.values()), '')
        severity = str(severity).upper()

        # Alternating row color
        row_fill = alt_row_fill if row % 2 == 0 else PatternFill(fill_type=None)

        for col, field in enumerate(fields, start=1):
            value = finding.get(field, '')
            
            if isinstance(value, list):
                if field == 'Resources':
                    # Format Resources field
                    formatted_resources = []
                    for resource in value:
                        resource_str = "Resource:\n"
                        for k, v in resource.items():
                            if k == 'Tags' and isinstance(v, dict):
                                resource_str += f"  {k}:\n"
                                for tag_k, tag_v in v.items():
                                    resource_str += f"    {tag_k}: {tag_v}\n"
                            else:
                                resource_str += f"  {k}: {v}\n"
                        formatted_resources.append(resource_str)
                    value = "\n\n".join(formatted_resources)
                else:
                    value = ', '.join(map(str, value))
            elif isinstance(value, dict):
                if field == 'Severity':
                    value = severity
                else:
                    value = ', '.join([f"{k}: {v}" for k, v in value.items()])
            
            if field == 'Description':
                # Split the description into bullet points using commas
                findings = value.split(', ')
                value = '\n'.join([f"• {finding.strip()}" for finding in findings if finding.strip()])
            
            cell = ws.cell(row=row, column=col, value=value)
            cell.fill = row_fill
            cell.alignment = Alignment(wrap_text=True, vertical='top')

            if field == 'Severity':
                if severity == 'CRITICAL':
                    cell.fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
                elif severity == 'HIGH':
                    cell.fill = PatternFill(start_color='FFFF9900', end_color='FFFF9900', fill_type='solid')
                elif severity == 'MEDIUM':
                    cell.fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
                elif severity == 'LOW':
                    cell.fill = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid')

    # Adjust column widths and add filters
    for col in range(1, len(fields) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 30
    ws.auto_filter.ref = ws.dimensions

    wb.save(output_file)

def main():
    parser = argparse.ArgumentParser(description='Process Prowler scan results.')
    parser.add_argument('-s', '--severity', type=str, default='chml',
                        help='Severity levels to include: c (Critical), h (High), m (Medium), l (Low). Default: chml')
    args = parser.parse_args()

    severity_map = {'c': 'CRITICAL', 'h': 'HIGH', 'm': 'MEDIUM', 'l': 'LOW'}
    severity_filter = [s.upper() for s in args.severity if s.lower() in severity_map]

    if not severity_filter:
        print("No valid severity levels selected. Please include at least one severity level (c, h, m, l).")
        return

    input_directory = 'input_scans'
    output_directory = 'output'
    
    # Ensure output directory exists
    os.makedirs(output_directory, exist_ok=True)
    
    # Get all JSON files in the input directory
    input_files = [os.path.join(input_directory, f) for f in os.listdir(input_directory) if f.endswith('.json')]
    
    if not input_files:
        print("No JSON files found in the input directory.")
        return
    
    compiled_data = process_prowler_scans(input_files, severity_filter)
    
    # Save JSON output
    json_output = os.path.join(output_directory, 'compiled_unique_findings.json')
    with open(json_output, 'w') as file:
        json.dump(compiled_data, file, indent=2)
    
    # Save Excel output
    excel_output = os.path.join(output_directory, 'compiled_unique_findings.xlsx')
    save_excel_file(compiled_data, excel_output)
    
    print(f"Processing complete. {len(compiled_data)} findings saved in the '{output_directory}' directory.")
    print(f"Severity levels included: {', '.join([severity_map[s.lower()] for s in args.severity if s.lower() in severity_map])}")

if __name__ == "__main__":
    main()
